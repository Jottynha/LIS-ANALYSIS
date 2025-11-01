"""
Análise de arquivos .lis para extrair tabelas; 
salvar em Excel com estatísticas;
gerar gráfico detalhado com ajuste gaussiano.
Dependências:
    pip install pandas openpyxl matplotlib numpy re argparse Path typing
"""

# Bibliotecas necessárias:
import re
from pathlib import Path
import argparse
from typing import Optional, Tuple, List, Dict
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------ Configurações e regex ------------------
START_MARKER = "The following is a distribution of peak overvoltages"
END_MARKER = "Summary of preceding table follows:"
STAT_TERMINATOR = "End of"
# regex para números (inteiros, floats, científicos); linguagem usada para definir padrões de busca em textos.
NUM_RE = re.compile(r'[-+]?(?:\d*\.\d+|\d+)(?:[eE][-+]?\d+)?')

# ---------- Parsing do .lis + extração de sumário ----------
def parse_lis_table(lis_path: Path) -> Tuple[Optional[pd.DataFrame], List[str], Dict[str, Tuple[Optional[float], Optional[float]]]]:
    """
    Lê o .lis, extrai a tabela de bins (colunas 6 números por linha),
    retorna (df, stats_lines_brutas, summary_dict)

    summary_dict (se encontrado) terá chaves: 'mean', 'variance', 'std_dev'
    e valores como tupla (grouped_value_or_None, ungrouped_value_or_None).
    """
    table_rows = []
    stats_lines: List[str] = []
    in_table = False

    with lis_path.open('r', errors='replace') as f:
        for raw_line in f:
            line = raw_line.rstrip('\n')
            if (not in_table) and (START_MARKER in line):
                in_table = True
                continue
            if in_table and (END_MARKER in line):
                in_table = False
                # coletar linhas de estatísticas brutas (texto) até linha vazia ou STAT_TERMINATOR
                for stat_raw in f:
                    stat_line = stat_raw.rstrip('\n')
                    if stat_line.strip() == "" or STAT_TERMINATOR in stat_line:
                        break
                    stats_lines.append(stat_line.replace(',', '.'))
                break
            if in_table:
                clean = line.replace(',', '.')
                nums = NUM_RE.findall(clean)
                # exige 6 números por linha
                if len(nums) >= 6:
                    try:
                        row_f = [float(x) for x in nums[:6]]
                        table_rows.append(row_f)
                    except ValueError:
                        continue

    summary = {}
    # tenta extrair Mean / Variance / Standard deviation das stats_lines
    for ln in stats_lines:
        low = ln.lower()
        if 'mean' in low:
            nums = NUM_RE.findall(ln.replace(',', '.'))
            # pode ter 1 ou 2 números; se 2 => (grouped, ungrouped)
            if len(nums) >= 2:
                summary['mean'] = (float(nums[0]), float(nums[1]))
            elif len(nums) == 1:
                summary['mean'] = (float(nums[0]), None)
        elif 'variance' in low:
            nums = NUM_RE.findall(ln.replace(',', '.'))
            if len(nums) >= 2:
                summary['variance'] = (float(nums[0]), float(nums[1]))
            elif len(nums) == 1:
                summary['variance'] = (float(nums[0]), None)
        elif 'standard deviation' in low or 'standard deviation' in ln.lower():
            nums = NUM_RE.findall(ln.replace(',', '.'))
            if len(nums) >= 2:
                summary['std_dev'] = (float(nums[0]), float(nums[1]))
            elif len(nums) == 1:
                summary['std_dev'] = (float(nums[0]), None)
    if not table_rows:
        return None, stats_lines, summary

    df = pd.DataFrame(table_rows, columns=[
        'Interval', 'Voltage_per_unit', 'Voltage_physical',
        'Frequency', 'Cumulative', 'Percent'
    ])

    # tentar converter colunas inteiras quando apropriado
    for col in ['Interval', 'Frequency', 'Cumulative']:
        try:
            if df[col].dropna().apply(float.is_integer).all() and not df[col].isna().any():
                df[col] = df[col].astype(int)
        except Exception:
            pass

    return df, stats_lines, summary

# ------------------ Salvar dados em Excel (aba 'Dados' e 'Estatisticas') ------------------

def save_df_to_excel_only(df: pd.DataFrame, out_path: Path, sheet_name: str = 'Dados'):
    """
    Salva o DataFrame na aba 'Dados' com formatação profissional:
    - Cabeçalhos com negrito e fundo azul claro
    - Autoajuste de colunas
    - Congelar painéis no cabeçalho
    - Filtros automáticos
    - Bordas nas células
    """
    mapping = {
        'Interval': 'Intervalo',
        'Voltage_per_unit': 'Tensao_pu',
        'Voltage_physical': 'Tensao_fisica',
        'Frequency': 'Frequencia',
        'Cumulative': 'Cumulativo',
        'Percent': 'Percentual'
    }
    df_to_save = df.rename(columns={k: v for k, v in mapping.items() if k in df.columns})

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)

    # Aplicar formatação profissional
    wb = load_workbook(out_path)
    ws = wb[sheet_name]
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Formatar cabeçalhos
    for i, col in enumerate(df_to_save.columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Autoajustar largura das colunas
        try:
            max_len = max(df_to_save[col].astype(str).map(len).max(), len(col)) + 2
        except Exception:
            max_len = len(col) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(max_len, 30)  # Máximo de 30
    
    # Formatar células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left")
    
    # Congelar painéis (primeira linha)
    ws.freeze_panes = ws['A2']
    
    # Adicionar filtros automáticos
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(out_path)
    wb.close()
    print(f"✅ Excel (aba '{sheet_name}') salvo com formatação profissional em: {out_path}")

# ------------------ Calcular estatísticas a partir dos bins (ponderadas) ------------------

def calcular_estatisticas_do_df(df: pd.DataFrame) -> dict:
    """
    Calcula estatísticas ponderadas a partir do DataFrame (usa Frequency quando disponível;
    pode derivar a partir de Cumulative/Percent se necessário).
    Retorna dict com as métricas numéricas.
    """
    cols = list(df.columns)
    voltage_candidates = ['Voltage_per_unit', 'Tensao_pu', 'voltagePerUnit', 'Voltage', 'Tensão', 'Tensao']
    freq_candidates = ['Frequency', 'Frequencia', 'Freq', 'Frequência']
    cumul_candidates = ['Cumulative', 'Cumulativo', 'Acumulado']
    percent_candidates = ['Percent', 'Percentual', 'Percent %', 'Percentagem']

    def _find(cands):
        for c in cands:
            for cc in cols:
                if str(cc).lower() == c.lower():
                    return cc
        return None

    voltage_col = _find(voltage_candidates)
    freq_col = _find(freq_candidates)
    cumul_col = _find(cumul_candidates)
    percent_col = _find(percent_candidates)

    df_num = df.copy()
    for c in df_num.columns:
        if df_num[c].dtype == object:
            df_num[c] = df_num[c].astype(str).str.replace(',', '.')
        df_num[c] = pd.to_numeric(df_num[c], errors='coerce')

    if voltage_col is None:
        for cand in ['Tensao_pu', 'Tensao', 'Tensão_pu', 'Tensão']:
            if cand in df_num.columns:
                voltage_col = cand
                break

    if voltage_col is None:
        raise ValueError("Coluna de tensão (pu) não encontrada no DataFrame.")

    voltage = df_num[voltage_col].to_numpy(dtype=float)

    # obter/derivar frequência
    freq = None
    method = None
    if freq_col is not None and df_num[freq_col].notna().any():
        freq = df_num[freq_col].fillna(0).to_numpy(dtype=float)
        method = 'freq_col'
    elif cumul_col is not None and df_num[cumul_col].notna().any():
        cumul = df_num[cumul_col].fillna(method='ffill').fillna(0).to_numpy(dtype=float)
        freq = np.diff(np.concatenate(([0.0], cumul)))
        method = 'derived_from_cumulative'
    elif percent_col is not None and df_num[percent_col].notna().any():
        pct = df_num[percent_col].fillna(0).to_numpy(dtype=float)
        if cumul_col is not None and not df_num[cumul_col].isna().all():
            total = float(df_num[cumul_col].dropna().iloc[-1])
        else:
            s = np.sum(pct)
            total = (s / 100.0) if s != 0 else 100.0
        freq = (pct / 100.0) * total
        method = 'derived_from_percent'
    else:
        for cc in df_num.columns:
            if 'cumul' in str(cc).lower():
                cumul = df_num[cc].fillna(method='ffill').fillna(0).to_numpy(dtype=float)
                freq = np.diff(np.concatenate(([0.0], cumul)))
                method = 'derived_from_cumulative_alt'
                break

    if freq is None:
        raise ValueError("Não foi possível detectar/derivar frequência (freq/cumul/percent).")

    n = min(len(voltage), len(freq))
    x = np.array(voltage[:n], dtype=float)
    y = np.array(freq[:n], dtype=float)
    mask = np.isfinite(x) & np.isfinite(y) & (y >= 0)
    x = x[mask]; y = y[mask]

    if x.size == 0 or y.size == 0 or np.sum(y) <= 0:
        raise ValueError("Dados insuficientes após limpeza para calcular estatísticas.")

    total_weight = float(np.sum(y))
    mu = float(np.sum(x * y) / total_weight)
    var = float(np.sum(y * (x - mu)**2) / total_weight)
    sigma = float(np.sqrt(var)) if var > 0 else 0.0

    cumsum = np.cumsum(y)
    median_val = float('nan')
    if total_weight > 0:
        idx_med = np.searchsorted(cumsum, total_weight / 2.0)
        median_val = float(x[idx_med]) if idx_med < len(x) else float(x[-1])

    mode_val = float(x[np.argmax(y)]) if y.size > 0 else float('nan')
    cv = float(sigma / mu) if mu != 0 else float('nan')

    if sigma > 0:
        skew = float(np.sum(y * (x - mu)**3) / (total_weight * sigma**3))
        kurt = float(np.sum(y * (x - mu)**4) / (total_weight * sigma**4) - 3.0)
    else:
        skew = float('nan')
        kurt = float('nan')

    # R² do ajuste gaussiano (escala por pico)
    if sigma > 0:
        pdf_x = np.exp(-0.5 * ((x - mu) / sigma)**2) / (sigma * np.sqrt(2 * np.pi))
        scale = (np.max(y) / np.max(pdf_x)) if np.max(pdf_x) > 0 else 1.0
        y_pred = pdf_x * scale
        ss_res = np.sum((y - y_pred)**2)
        ss_tot = np.sum((y - np.mean(y))**2)
        r2 = float(1.0 - ss_res / ss_tot) if ss_tot > 0 else float('nan')
    else:
        r2 = float('nan')

    stats = {
        'mean': mu,
        'variance': var,
        'std_dev': sigma,
        'median': median_val,
        'mode': mode_val,
        'total_freq': total_weight,
        'cv': cv,
        'skewness': skew,
        'kurtosis': kurt,
        'r2': r2,
        'freq_method': method
    }
    return stats

# ------------------ Escrever Estátisticas no Excel ------------------

def escrever_estatisticas_excel(excel_path: Path, computed_stats: dict,
                                summary_from_lis: Dict[str, Tuple[Optional[float], Optional[float]]] = None,
                                sheet_name: str = 'Estatisticas'):
    """
    Escreve estatísticas com formatação profissional em layout compacto lado a lado:
    - Tabela do .lis (se houver) à esquerda
    - Estatísticas computadas à direita
    - Formatação com cores, negrito e bordas
    """
    if not excel_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {excel_path}")

    wb = load_workbook(excel_path)
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(size=10)
    
    title_font = Font(bold=True, size=12, color="2F75B5")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")

    row = 1
    
    # SEÇÃO 1: Estatísticas do .lis (se existirem) - COLUNA A-C
    if summary_from_lis:
        # Título
        cell = ws.cell(row=row, column=1, value='📊 Estatísticas do Arquivo .lis')
        cell.font = title_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        
        # Cabeçalhos
        for col, text in [(1, 'Métrica'), (2, 'Grouped'), (3, 'Ungrouped')]:
            cell = ws.cell(row=row, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        row += 1
        
        # Dados
        for key, pretty in [('mean', 'Média'), ('variance', 'Variância'), ('std_dev', 'Desvio Padrão')]:
            if key in summary_from_lis:
                g, u = summary_from_lis.get(key, (None, None))
                
                # Métrica
                cell = ws.cell(row=row, column=1, value=pretty)
                cell.font = data_font
                cell.alignment = left_alignment
                cell.border = thin_border
                
                # Grouped
                cell = ws.cell(row=row, column=2)
                if g is not None:
                    cell.value = float(g)
                    cell.number_format = '0.000000E+00'
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Ungrouped
                cell = ws.cell(row=row, column=3)
                if u is not None:
                    cell.value = float(u)
                    cell.number_format = '0.000000E+00'
                cell.font = data_font
                cell.alignment = center_alignment
                cell.border = thin_border
                
                row += 1
        
        row += 1  # Espaço

    # SEÇÃO 2: Estatísticas Computadas - LAYOUT COMPACTO (2 COLUNAS)
    start_row_computed = row
    
    # Título
    cell = ws.cell(row=row, column=1, value='🔬 Estatísticas Computadas')
    cell.font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    
    # Cabeçalhos
    for col, text in [(1, 'Métrica'), (2, 'Valor')]:
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    # Preparar dados
    keys_order = [
        'mean', 'variance', 'std_dev', 'median', 'mode',
        'total_freq', 'cv', 'skewness', 'kurtosis', 'r2', 'freq_method'
    ]
    pretty_names = {
        'mean': 'Média (μ)',
        'variance': 'Variância (σ²)',
        'std_dev': 'Desvio Padrão (σ)',
        'median': 'Mediana',
        'mode': 'Moda',
        'total_freq': 'Σ Frequências',
        'cv': 'Coef. Variação (CV)',
        'skewness': 'Assimetria',
        'kurtosis': 'Curtose',
        'r2': 'R² Ajuste',
        'freq_method': 'Método'
    }
    number_formats = {
        'mean': '0.000000',
        'variance': '0.000000E+00',
        'std_dev': '0.000000',
        'median': '0.000000',
        'mode': '0.000000',
        'total_freq': '0',
        'cv': '0.0000',
        'skewness': '0.0000',
        'kurtosis': '0.0000',
        'r2': '0.0000'
    }
    
    present_keys = [k for k in keys_order if k in computed_stats]
    
    # Escrever dados
    for key in present_keys:
        # Métrica
        cell = ws.cell(row=row, column=1, value=pretty_names.get(key, key))
        cell.font = data_font
        cell.alignment = left_alignment
        cell.border = thin_border
        
        # Valor
        cell = ws.cell(row=row, column=2)
        val = computed_stats.get(key)
        if isinstance(val, (int, float)) and not (isinstance(val, float) and np.isnan(val)):
            cell.value = float(val)
            fmt = number_formats.get(key)
            if fmt:
                cell.number_format = fmt
        else:
            cell.value = str(val) if val else '-'
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        
        row += 1
    
    # Ajustar larguras das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    
    # Congelar painéis
    ws.freeze_panes = ws['A3'] if summary_from_lis else ws['A' + str(start_row_computed + 2)]
    
    wb.save(excel_path)
    wb.close()
    print(f"✅ Estatísticas salvas com formatação profissional na aba '{sheet_name}'")

# ------------------ Função do gráfico (lê o Excel gerado) ------------------

def criar_grafico_a_partir_do_excel(excel_path: Path, outdir: Path, sim_index: int = 1,
                                    salvar_png: bool = True, mostrar: bool = False) -> Optional[Path]:
    """
    Lê o Excel em `excel_path` (aba 'Dados'), obtém tensão/frequência e plota gráfico detalhado.
    Retorna Path do PNG ou None.
    """
    if not excel_path.exists():
        print("Arquivo Excel não encontrado:", excel_path)
        return None

    try:
        df_excel = pd.read_excel(excel_path, sheet_name='Dados')
    except Exception as e:
        print("Erro ao ler o Excel:", e)
        return None

    # detecta colunas candidatas
    def _find_column(candidates, cols):
        for c in candidates:
            for cc in cols:
                if c.lower() == str(cc).lower():
                    return cc
        return None

    cols = list(df_excel.columns)
    voltage_candidates = ['Voltage_per_unit', 'Tensao_pu', 'voltagePerUnit', 'Tensão_pu', 'Tensão (pu)', 'Tensao', 'Voltage']
    freq_candidates = ['Frequency', 'Frequencia', 'Freq', 'Frequência']
    cumul_candidates = ['Cumulative', 'Cumulativo', 'CumulativeCount', 'Acumulado']
    percent_candidates = ['Percent', 'Percentual', 'Percent %', 'Percentagem']

    voltage_col = _find_column(voltage_candidates, cols)
    freq_col = _find_column(freq_candidates, cols)
    cumul_col = _find_column(cumul_candidates, cols)
    percent_col = _find_column(percent_candidates, cols)

    if voltage_col is None:
        for cand in ['Tensao_pu', 'Tensao', 'Tensão_pu', 'Tensão']:
            if cand in cols:
                voltage_col = cand
                break

    df_num = df_excel.copy()
    for c in df_num.columns:
        if df_num[c].dtype == object:
            df_num[c] = df_num[c].astype(str).str.replace(',', '.')
        df_num[c] = pd.to_numeric(df_num[c], errors='coerce')

    if voltage_col is None:
        print("Não encontrei coluna de tensão (pu) no Excel. Colunas:", cols)
        return None

    voltage_series = df_num[voltage_col]

    # obter freq
    freq_series = None
    method_used = None
    if freq_col is not None and df_num[freq_col].notna().any():
        freq_series = df_num[freq_col].fillna(0)
        method_used = "freq_col"
    elif cumul_col is not None and df_num[cumul_col].notna().any():
        cumul = df_num[cumul_col].fillna(method='ffill').fillna(0).to_numpy(dtype=float)
        freq = np.diff(np.concatenate(([0.0], cumul)))
        freq_series = pd.Series(freq)
        method_used = "derived_from_cumulative"
    elif percent_col is not None and df_num[percent_col].notna().any():
        pct = df_num[percent_col].fillna(0).to_numpy(dtype=float)
        total = None
        if cumul_col is not None and not df_num[cumul_col].isna().all():
            total = float(df_num[cumul_col].dropna().iloc[-1])
        else:
            s = np.sum(pct)
            total = (s / 100.0) if s != 0 else 100.0
        freq = (pct / 100.0) * total
        freq_series = pd.Series(freq)
        method_used = "derived_from_percent"
    else:
        for cc in df_num.columns:
            if 'cumul' in str(cc).lower():
                cumul_col = cc
                cumul = df_num[cumul_col].fillna(method='ffill').fillna(0).to_numpy(dtype=float)
                freq = np.diff(np.concatenate(([0.0], cumul)))
                freq_series = pd.Series(freq)
                method_used = "derived_from_cumulative_alt"
                break

    if freq_series is None:
        print("Não foi possível determinar frequências a partir do Excel.")
        return None

    # alinhar, limpar e ordenar
    x = voltage_series.to_numpy(dtype=float)
    y = freq_series.to_numpy(dtype=float)
    n = min(len(x), len(y))
    x = x[:n]; y = y[:n]
    mask = np.isfinite(x) & np.isfinite(y) & (y >= 0)
    x = x[mask]; y = y[mask]
    if x.size == 0 or y.size == 0 or np.sum(y) <= 0:
        print("Dados insuficientes após limpeza.")
        return None

    order = np.argsort(x)
    x = x[order]; y = y[order]
    total_weight = np.sum(y)

    # calcular estatísticas (novamente) e escrever formatadas (se quiser sobrescrever)
    try:
        computed_stats = calcular_estatisticas_do_df(df_excel)
    except Exception:
        computed_stats = {}

    # tenta ler summary do excel (se foi gravado)
    summary_from_excel = {}
    try:
        wb = load_workbook(excel_path, data_only=True)
        if 'Estatisticas' in wb.sheetnames:
            ws = wb['Estatisticas']
            # Tentativa simples: ler as células da tabela Grouped/Ungrouped no topo (se existirem)
            # procuramos por 'Mean' ou 'Mean (do .lis)' na coluna A
            for r in range(1, 10):
                a = ws.cell(row=r, column=1).value
                if isinstance(a, str) and 'mean' in a.lower():
                    g = ws.cell(row=r, column=2).value
                    u = ws.cell(row=r, column=3).value
                    if g is not None or u is not None:
                        summary_from_excel['mean'] = (float(g) if g is not None else None, float(u) if u is not None else None)
                if isinstance(a, str) and 'variance' in a.lower():
                    g = ws.cell(row=r, column=2).value
                    u = ws.cell(row=r, column=3).value
                    if g is not None or u is not None:
                        summary_from_excel['variance'] = (float(g) if g is not None else None, float(u) if u is not None else None)
                if isinstance(a, str) and 'standard' in a.lower():
                    g = ws.cell(row=r, column=2).value
                    u = ws.cell(row=r, column=3).value
                    if g is not None or u is not None:
                        summary_from_excel['std_dev'] = (float(g) if g is not None else None, float(u) if u is not None else None)
        wb.close()
    except Exception:
        pass
    mu = computed_stats.get('mean') if 'mean' in computed_stats else (summary_from_excel.get('mean', (None, None))[0] if 'mean' in summary_from_excel else np.nan)
    sigma = computed_stats.get('std_dev') if 'std_dev' in computed_stats else (summary_from_excel.get('std_dev', (None, None))[0] if 'std_dev' in summary_from_excel else np.nan)

    # gerar curva gaussiana
    x_smooth = np.linspace(np.min(x), np.max(x), 800)
    if sigma and np.isfinite(sigma) and sigma > 0:
        pdf = np.exp(-0.5 * ((x_smooth - mu) / sigma)**2) / (sigma * np.sqrt(2 * np.pi))
        scale_factor = (np.max(y) / np.max(pdf)) if np.max(pdf) > 0 else 1.0
        y_smooth = pdf * scale_factor
        pdf_x = np.exp(-0.5 * ((x - mu) / sigma)**2) / (sigma * np.sqrt(2 * np.pi))
        y_pred_x = pdf_x * scale_factor
    else:
        y_smooth = np.zeros_like(x_smooth)
        y_pred_x = np.zeros_like(x)

    # R²
    ss_res = np.sum((y - y_pred_x)**2)
    ss_tot = np.sum((y - np.mean(y))**2)
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan

    # estimativa de largura de barra
    unique_x = np.unique(x)
    if unique_x.size > 1:
        diffs = np.diff(unique_x)
        diffs_pos = diffs[diffs > 0]
        bin_width = float(np.median(diffs_pos)) if diffs_pos.size > 0 else (np.max(x) - np.min(x)) / max(1, len(unique_x))
    else:
        bin_width = 0.1 if unique_x.size == 1 else 1.0
    bar_width = bin_width * 0.9

    # plot
    fig, ax = plt.subplots(figsize=(11, 7))
    ax.bar(x, y, width=bar_width, alpha=0.35, label='Frequência (bins)', align='center', edgecolor='k', linewidth=0.3)
    ax.scatter(x, y, color='tab:blue', s=30, zorder=5, label='Pontos (x vs freq)')
    if np.any(y_smooth):
        ax.plot(x_smooth, y_smooth, color='tab:orange', linewidth=2.2, label='Ajuste Gaussiano')

    ax.set_xlabel('Tensão (pu)')
    ax.set_ylabel('Frequência')
    ax.grid(alpha=0.25)
    ax.legend(loc='upper left')

    # eixo secundário acumulado %
    ax2 = ax.twinx()
    cumsum = np.cumsum(y)
    cum_pct = (cumsum / total_weight) * 100.0
    ax2.plot(x, cum_pct, color='tab:green', marker='o', linestyle='--', label='Acumulado (%)')
    ax2.set_ylabel('Acumulado (%)')
    ax2.set_ylim(0, 100)

    # caixa de estatísticas (resumo visual)
    pretty_stats_text = (
        f"μ = {mu:.6g}\n"
        f"σ = {sigma:.6g}\n"
        f"Mediana = {computed_stats.get('median', float('nan')):.6g}\n"
        f"Moda = {computed_stats.get('mode', float('nan')):.6g}\n"
        f"Soma freq = {computed_stats.get('total_freq', float('nan')):.6g}\n"
        f"CV = {computed_stats.get('cv', float('nan')):.6g}\n"
        f"Skewness = {computed_stats.get('skewness', float('nan')):.6g}\n"
        f"Kurtosis = {computed_stats.get('kurtosis', float('nan')):.6g}\n"
        f"R² = {r2:.5g}\n"
        f"Método freq = {computed_stats.get('freq_method')}"
    )
    bbox_props = dict(boxstyle="round,pad=0.6", fc="white", ec="0.4", alpha=0.9)
    ax.text(0.98, 0.95, pretty_stats_text, transform=ax.transAxes, fontsize=9,
            verticalalignment='top', horizontalalignment='right', bbox=bbox_props)

    ax.set_title(f"Ajuste Gaussiano Detalhado — {excel_path.stem}  (sim {sim_index})")

    lines, labels = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax2.legend(lines + lines2, labels + labels2, loc='lower right')

    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    out_png = outdir / f"gauss_detalhado_{sim_index}.png"
    try:
        plt.tight_layout()
        if salvar_png:
            plt.savefig(out_png, dpi=220, bbox_inches='tight')
            print("Gráfico detalhado salvo em:", out_png)
        if mostrar:
            plt.show()
    finally:
        plt.close(fig)

    return out_png

# ------------------ Seleção interativa e helpers para múltiplos arquivos ------------------

def _parse_indices_input(s: str, max_n: int) -> List[int]:
    """Converte string tipo '1,3-5' em lista de índices (1-based) válidos até max_n."""
    s = (s or "").strip()
    if not s:
        return []
    sel: set = set()
    for part in s.split(','):
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            try:
                a_str, b_str = part.split('-', 1)
                a = int(a_str)
                b = int(b_str)
                if a > b:
                    a, b = b, a
                for i in range(a, b + 1):
                    if 1 <= i <= max_n:
                        sel.add(i)
            except Exception:
                continue
        else:
            try:
                i = int(part)
                if 1 <= i <= max_n:
                    sel.add(i)
            except Exception:
                continue
    return sorted(sel)


def selecionar_arquivos_interativo(folder: Path) -> List[Path]:
    """Lista arquivos .lis na pasta e permite seleção múltipla via input."""
    files = sorted(folder.glob('*.lis'), key=lambda f: f.stat().st_mtime, reverse=True)
    if not files:
        print("Nenhum arquivo .lis encontrado na pasta:", folder)
        return []
    print("Arquivos .lis encontrados (mais recentes primeiro):")
    for idx, f in enumerate(files, start=1):
        try:
            t = f.stat().st_mtime
        except Exception:
            t = 0
        print(f"  {idx:>2d}) {f.name}")
    print("Digite os índices desejados (ex: 1,3-5) e pressione Enter. Deixe vazio para cancelar.")
    choice = input("> ").strip()
    idxs = _parse_indices_input(choice, len(files))
    if not idxs:
        print("Nenhuma seleção realizada.")
        return []
    selected = [files[i - 1] for i in idxs]
    print("Selecionados:", ', '.join([p.name for p in selected]))
    return selected


def obter_xy_e_stats_de_excel(excel_path: Path):
    """Extrai (x,y) e (mu,sigma) do Excel gerado (aba 'Dados'). Retorna (x, y, mu, sigma) ou None em falha."""
    if not excel_path.exists():
        return None
    try:
        df_excel = pd.read_excel(excel_path, sheet_name='Dados')
    except Exception:
        return None

    def _find_column(candidates, cols):
        for c in candidates:
            for cc in cols:
                if c.lower() == str(cc).lower():
                    return cc
        return None

    cols = list(df_excel.columns)
    voltage_candidates = ['Voltage_per_unit', 'Tensao_pu', 'voltagePerUnit', 'Tensão_pu', 'Tensão (pu)', 'Tensao', 'Voltage']
    freq_candidates = ['Frequency', 'Frequencia', 'Freq', 'Frequência']
    cumul_candidates = ['Cumulative', 'Cumulativo', 'CumulativeCount', 'Acumulado']
    percent_candidates = ['Percent', 'Percentual', 'Percent %', 'Percentagem']

    voltage_col = _find_column(voltage_candidates, cols)
    freq_col = _find_column(freq_candidates, cols)
    cumul_col = _find_column(cumul_candidates, cols)
    percent_col = _find_column(percent_candidates, cols)

    if voltage_col is None:
        for cand in ['Tensao_pu', 'Tensao', 'Tensão_pu', 'Tensão']:
            if cand in cols:
                voltage_col = cand
                break

    df_num = df_excel.copy()
    for c in df_num.columns:
        if df_num[c].dtype == object:
            df_num[c] = df_num[c].astype(str).str.replace(',', '.')
        df_num[c] = pd.to_numeric(df_num[c], errors='coerce')

    if voltage_col is None:
        return None

    voltage_series = df_num[voltage_col]

    # obter freq
    freq_series = None
    if freq_col is not None and df_num[freq_col].notna().any():
        freq_series = df_num[freq_col].fillna(0)
    elif cumul_col is not None and df_num[cumul_col].notna().any():
        cumul = df_num[cumul_col].fillna(method='ffill').fillna(0).to_numpy(dtype=float)
        freq = np.diff(np.concatenate(([0.0], cumul)))
        freq_series = pd.Series(freq)
    elif percent_col is not None and df_num[percent_col].notna().any():
        pct = df_num[percent_col].fillna(0).to_numpy(dtype=float)
        total = None
        if cumul_col is not None and not df_num[cumul_col].isna().all():
            total = float(df_num[cumul_col].dropna().iloc[-1])
        else:
            s = np.sum(pct)
            total = (s / 100.0) if s != 0 else 100.0
        freq = (pct / 100.0) * total
        freq_series = pd.Series(freq)
    else:
        for cc in df_num.columns:
            if 'cumul' in str(cc).lower():
                cumul = df_num[cc].fillna(method='ffill').fillna(0).to_numpy(dtype=float)
                freq = np.diff(np.concatenate(([0.0], cumul)))
                freq_series = pd.Series(freq)
                break

    if freq_series is None:
        return None

    x = voltage_series.to_numpy(dtype=float)
    y = freq_series.to_numpy(dtype=float)
    n = min(len(x), len(y))
    x = x[:n]; y = y[:n]
    mask = np.isfinite(x) & np.isfinite(y) & (y >= 0)
    x = x[mask]; y = y[mask]
    if x.size == 0 or y.size == 0 or np.sum(y) <= 0:
        return None

    order = np.argsort(x)
    x = x[order]
    y = y[order]

    # stats
    try:
        computed_stats = calcular_estatisticas_do_df(df_excel)
    except Exception:
        computed_stats = {}

    mu = computed_stats.get('mean', np.nan)
    sigma = computed_stats.get('std_dev', np.nan)

    return x, y, mu, sigma


def criar_grafico_comparativo(excel_paths: List[Path], outdir: Path, mostrar: bool = False) -> Optional[Path]:
    """Gera gráfico comparativo sobrepondo séries e ajustes gaussianos de múltiplos Excel gerados."""
    series = []
    labels = []
    for p in excel_paths:
        res = obter_xy_e_stats_de_excel(p)
        if res is None:
            print("Aviso: não foi possível extrair dados de:", p)
            continue
        x, y, mu, sigma = res
        series.append((x, y, mu, sigma))
        labels.append(p.stem)

    if not series:
        print("Sem dados para gráfico comparativo.")
        return None

    fig, ax = plt.subplots(figsize=(12, 7))
    for (x, y, mu, sigma), label in zip(series, labels):
        # pontos (suavemente)
        ax.scatter(x, y, s=20, alpha=0.5, label=f"{label} pontos")
        # ajuste gaussiano
        if sigma and np.isfinite(sigma) and sigma > 0:
            x_smooth = np.linspace(np.min(x), np.max(x), 800)
            pdf = np.exp(-0.5 * ((x_smooth - mu) / sigma)**2) / (sigma * np.sqrt(2 * np.pi))
            scale_factor = (np.max(y) / np.max(pdf)) if np.max(pdf) > 0 else 1.0
            y_smooth = pdf * scale_factor
            ax.plot(x_smooth, y_smooth, linewidth=2.0, label=f"{label} ajuste")

    ax.set_xlabel('Tensão (pu)')
    ax.set_ylabel('Frequência')
    ax.grid(alpha=0.25)
    ax.set_title('Comparativo — Distribuição e Ajuste Gaussiano')
    ax.legend(ncol=2, fontsize=8)

    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    out_png = outdir / "gauss_comparativo.png"
    try:
        plt.tight_layout()
        plt.savefig(out_png, dpi=220, bbox_inches='tight')
        print("Gráfico comparativo salvo em:", out_png)
        if mostrar:
            plt.show()
    finally:
        plt.close(fig)
    return out_png

# ------------------ Fluxo principal ------------------

def main():
    parser = argparse.ArgumentParser(description="Analisa arquivos .lis e gera Excel + gráfico (com opção de comparativo).")
    parser.add_argument('--folder', default='.', help='Pasta para procurar o .lis (padrão = atual).')
    parser.add_argument('--sim-index', type=int, default=1, help='Índice inicial para nomear arquivos de saída.')
    parser.add_argument('--outdir', default='Simulation_Result', help='Pasta de saída.')
    parser.add_argument('--select', action='store_true', help='Abrir seleção interativa de arquivos .lis (multi-seleção).')
    parser.add_argument('--lis', nargs='*', help='Lista de arquivos .lis para processar (pode múltiplos).')
    parser.add_argument('--gui', action='store_true', help='Abrir interface gráfica (Tkinter).')
    args = parser.parse_args()

    # GUI override
    if args.gui:
        try:
            from gui import launch_gui
        except Exception as e:
            print('Erro ao carregar GUI:', e)
            raise SystemExit(2)
        launch_gui(Path(args.folder).resolve(), Path(args.outdir).resolve(), args.sim_index)
        return

    folder = Path(args.folder).resolve()
    outdir = Path(args.outdir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    # Descobrir arquivos a processar
    selected_files: List[Path] = []
    if args.lis:
        for item in args.lis:
            p = Path(item)
            if not p.is_absolute():
                p = (folder / p).resolve()
            if p.exists() and p.suffix.lower() == '.lis':
                selected_files.append(p)
            else:
                print("Ignorando (não encontrado ou extensão diferente de .lis):", item)
    elif args.select:
        selected_files = selecionar_arquivos_interativo(folder)
    else:
        lis_files = list(folder.glob('*.lis'))
        if not lis_files:
            print("Nenhum arquivo .lis encontrado na pasta:", folder)
            raise SystemExit(1)
        lis_path = max(lis_files, key=lambda f: f.stat().st_mtime)
        selected_files = [lis_path]

    if not selected_files:
        print("Nada a processar.")
        raise SystemExit(0)

    excel_paths: List[Path] = []
    for idx, lis_path in enumerate(selected_files, start=args.sim_index):
        print("Usando .lis:", lis_path)
        # parse do .lis
        df, stats_lines, summary_from_lis = parse_lis_table(lis_path)
        if df is None:
            print("Tabela não encontrada no .lis (nenhuma linha com 6 números detectada):", lis_path)
            continue

        # salva aba 'Dados'
        excel_path = outdir / f"Resultados_Simulacao_{idx}.xlsx"
        save_df_to_excel_only(df, excel_path)

        # calcula estatísticas ponderadas a partir dos bins
        try:
            computed_stats = calcular_estatisticas_do_df(df)
        except Exception as e:
            print("Erro ao calcular estatísticas a partir dos bins:", e)
            computed_stats = {}

        # escreve a aba 'Estatisticas' (inclui os valores extraídos do .lis, se houver)
        try:
            escrever_estatisticas_excel(excel_path, computed_stats, summary_from_lis=summary_from_lis)
        except Exception as e:
            print("Falha ao escrever estatísticas no Excel:", e)
            # fallback: tenta salvar CSV
            try:
                csv_path = outdir / f"estatisticas_sim_{idx}.csv"
                df_csv = pd.DataFrame([computed_stats])
                df_csv.to_csv(csv_path, index=False)
                print("Fallback: estatísticas salvas em CSV:", csv_path)
            except Exception:
                pass

        # criar o gráfico com base no Excel gerado
        print("Criando gráfico individual...")
        _ = criar_grafico_a_partir_do_excel(excel_path, outdir, sim_index=idx, salvar_png=True, mostrar=False)
        excel_paths.append(excel_path)

    # Se houver múltiplos, cria gráfico comparativo sobreposto
    if len(excel_paths) > 1:
        print("Gerando gráfico comparativo sobreposto...")
        _ = criar_grafico_comparativo(excel_paths, outdir, mostrar=False)

    print("Processo concluído. Verifique a pasta:", outdir)

if __name__ == "__main__":
    main()
