from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from lis_analysis.main import (
    calcular_estatisticas_do_df,
    escrever_estatisticas_excel,
    parse_lis_once,
    save_df_to_excel_only,
    save_time_series_to_excel,
)


GOLDEN_CASES = [
    {
        "name": "convencional_sem_controle",
        "path": Path("data/samples/Listas ATP/CONVENCIONAL/Caso0_ReEnergizacao_Convenc_SemControle.LIS"),
        "table_shape": (22, 6),
        "ts_shape": (2001, 9),
        "grouped_mean": 2.22283333,
        "grouped_variance": 0.0203046544,
        "grouped_std_dev": 0.142494401,
        "computed_mean": 2.2228333333333334,
        "computed_variance": 0.020304654403567452,
        "computed_std_dev": 0.14249440130604238,
        "median": 2.225,
        "mode": 2.275,
        "total_freq": 300.0,
        "first_row": {
            "Interval": 32,
            "Voltage_per_unit": 1.60,
            "Frequency": 0.0,
            "Percent": 100.0,
        },
    },
    {
        "name": "convencional_rpi_500",
        "path": Path("data/samples/Listas ATP/CONVENCIONAL/Caso0_ReEnergizacao_Convenc_RPI=500 e RF=30.lis"),
        "table_shape": (9, 6),
        "ts_shape": (2001, 9),
        "grouped_mean": 1.6805,
        "grouped_variance": 0.00161680602,
        "grouped_std_dev": 0.0402095265,
        "computed_mean": 1.6804999999999994,
        "computed_variance": 0.001616806020066892,
        "computed_std_dev": 0.040209526483992473,
        "median": 1.675,
        "mode": 1.675,
        "total_freq": 300.0,
        "first_row": {
            "Interval": 33,
            "Voltage_per_unit": 1.65,
            "Frequency": 0.0,
            "Percent": 100.0,
        },
    },
    {
        "name": "otimizada_rpi_500",
        "path": Path("data/samples/Listas ATP/OTIMIZADA/Caso0_ReEnergizacao_OTIMIZADA_RPI=500 e RF=30.lis"),
        "table_shape": (9, 6),
        "ts_shape": (2001, 9),
        "grouped_mean": 1.68083333,
        "grouped_variance": 0.0018136845,
        "grouped_std_dev": 0.0425873749,
        "computed_mean": 1.680833333333333,
        "computed_variance": 0.0018136845039018969,
        "computed_std_dev": 0.04258737493555921,
        "median": 1.675,
        "mode": 1.675,
        "total_freq": 300.0,
        "first_row": {
            "Interval": 33,
            "Voltage_per_unit": 1.65,
            "Frequency": 0.0,
            "Percent": 100.0,
        },
    },
]


class AcceptanceGoldenCasesTest(unittest.TestCase):
    def _find_rows_by_label(self, ws, label: str) -> list[int]:
        return [row for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == label]

    def test_golden_cases_match_reference_analysis(self):
        for case in GOLDEN_CASES:
            with self.subTest(case=case["name"]):
                parsed = parse_lis_once(case["path"], verbose=False)
                stats = calcular_estatisticas_do_df(parsed.table_df)

                self.assertIsNotNone(parsed.table_df)
                self.assertIsNotNone(parsed.time_series_df)
                self.assertEqual(tuple(parsed.table_df.shape), case["table_shape"])
                self.assertEqual(tuple(parsed.time_series_df.shape), case["ts_shape"])

                first_row = parsed.table_df.iloc[0]
                for key, expected in case["first_row"].items():
                    self.assertAlmostEqual(float(first_row[key]), float(expected), places=6)

                self.assertAlmostEqual(float(parsed.summary["mean"][0]), case["grouped_mean"], places=6)
                self.assertAlmostEqual(float(parsed.summary["variance"][0]), case["grouped_variance"], places=6)
                self.assertAlmostEqual(float(parsed.summary["std_dev"][0]), case["grouped_std_dev"], places=6)

                self.assertAlmostEqual(float(stats["mean"]), case["computed_mean"], places=6)
                self.assertAlmostEqual(float(stats["variance"]), case["computed_variance"], places=6)
                self.assertAlmostEqual(float(stats["std_dev"]), case["computed_std_dev"], places=6)
                self.assertAlmostEqual(float(stats["median"]), case["median"], places=6)
                self.assertAlmostEqual(float(stats["mode"]), case["mode"], places=6)
                self.assertAlmostEqual(float(stats["total_freq"]), case["total_freq"], places=6)
                self.assertAlmostEqual(float(stats["bin_width"]), 0.05, places=6)

    def test_golden_case_exports_keep_expected_workbook_content(self):
        case = GOLDEN_CASES[0]
        parsed = parse_lis_once(case["path"], verbose=False)
        stats = calcular_estatisticas_do_df(parsed.table_df)

        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "golden_case.xlsx"
            save_df_to_excel_only(parsed.table_df, out_path)
            escrever_estatisticas_excel(out_path, stats, summary_from_lis=parsed.summary)
            save_time_series_to_excel(parsed.time_series_df, out_path)

            wb = load_workbook(out_path, data_only=True)
            self.assertIn("Dados", wb.sheetnames)
            self.assertIn("Estatisticas", wb.sheetnames)
            self.assertIn("Dados_Temporais", wb.sheetnames)

            ws_dados = wb["Dados"]
            ws_stats = wb["Estatisticas"]
            ws_ts = wb["Dados_Temporais"]

            self.assertEqual(ws_dados["A2"].value, case["first_row"]["Interval"])
            self.assertAlmostEqual(float(ws_dados["B2"].value), case["first_row"]["Voltage_per_unit"], places=6)
            self.assertEqual(ws_ts["A1"].value, "Step")
            self.assertEqual(ws_ts["B1"].value, "Time")
            self.assertEqual(ws_ts.max_row, case["ts_shape"][0] + 1)

            grouped_media_row = self._find_rows_by_label(ws_stats, "Média")[0]
            grouped_variance_row = self._find_rows_by_label(ws_stats, "Variância")[0]
            computed_media_row = self._find_rows_by_label(ws_stats, "Média (μ)")[0]
            computed_variance_row = self._find_rows_by_label(ws_stats, "Variância (σ²)")[0]

            self.assertAlmostEqual(float(ws_stats.cell(row=grouped_media_row, column=2).value), case["grouped_mean"], places=6)
            self.assertAlmostEqual(float(ws_stats.cell(row=grouped_variance_row, column=2).value), case["grouped_variance"], places=6)
            self.assertAlmostEqual(float(ws_stats.cell(row=computed_media_row, column=2).value), case["computed_mean"], places=6)
            self.assertAlmostEqual(float(ws_stats.cell(row=computed_variance_row, column=2).value), case["computed_variance"], places=6)

            wb.close()


if __name__ == "__main__":
    unittest.main()
